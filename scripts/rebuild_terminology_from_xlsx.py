#!/usr/bin/env python3
import json
import re
from collections import OrderedDict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC_XLSX = ROOT / "source" / "reference" / "疑问汇总及术语对照表.xlsx"
OUT_MD = ROOT / "术语对照表整理.md"
OUT_JSON = ROOT / "data" / "terminology.json"

SHEET = "综合术语表"
SECTION1_HEADER_ROW = 3
SECTION1_START_ROW = 4
SECTION2_HEADER_ROW = 462
SECTION2_START_ROW = 463

CASE_PREFIXES = (
    "Ablative",
    "Accusative",
    "Dative",
    "Genitive",
    "Nominative",
    "Vocative",
    "Locative",
)


def clean_text(v):
    if v is None:
        return ""
    t = str(v).replace("\xa0", " ").strip()
    t = re.sub(r"\s+", " ", t)
    return t


def normalize_term(t):
    t = clean_text(t).lower()
    t = t.replace("instrunent", "instrument")
    t = re.sub(r"[，。；;:,.]+$", "", t)
    t = re.sub(r"\s+", " ", t)
    return t


def slugify(text):
    s = normalize_term(text)
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "term"


def extract_english_from_mixed(text):
    t = clean_text(text)
    if not t:
        return ""
    # If CJK exists, keep trailing Latin segment as the term.
    if re.search(r"[\u4e00-\u9fff]", t):
        m = re.search(r"([A-Za-zÀ-ÖØ-öø-ÿĀ-ž][A-Za-zÀ-ÖØ-öø-ÿĀ-ž0-9 \-_/(),.'’]*)$", t)
        if m:
            return clean_text(m.group(1))
    return t


def maybe_restore_prefix(term, current_prefix):
    t = clean_text(term)
    if not t:
        return t, current_prefix

    t = re.sub(r"[，。；;:,.]+$", "", t)
    low = t.lower()
    if low.startswith("of ") and current_prefix:
        t = f"{current_prefix} {t}"
    elif low.startswith("with ") and current_prefix:
        t = f"{current_prefix} {t}"

    for prefix in CASE_PREFIXES:
        if re.match(rf"^{re.escape(prefix)}\b", t, flags=re.IGNORECASE):
            t = prefix + t[len(prefix):]
            current_prefix = prefix
            break

    return t, current_prefix


def build_academic_map(ws):
    out = {}
    for r in range(SECTION2_START_ROW, ws.max_row + 1):
        col_b = clean_text(ws.cell(r, 2).value)
        col_c = clean_text(ws.cell(r, 3).value)
        if not col_b and not col_c:
            continue
        if "【" in col_b or "【" in col_c:
            continue
        if len(col_b) == 1 and col_b.isalpha() and col_b.upper() == col_b and col_b == col_c:
            continue
        term = extract_english_from_mixed(col_b)
        if not term or not re.search(r"[A-Za-zÀ-ÖØ-öø-ÿĀ-ž]", term):
            continue
        if not col_c:
            continue
        out[normalize_term(term)] = col_c
    return out


def parse_section1_rows(ws, academic_map):
    rows = []
    current_prefix = ""

    for r in range(SECTION1_START_ROW, SECTION2_HEADER_ROW):
        raw = [clean_text(ws.cell(r, c).value) for c in range(1, 7)]
        if not any(raw):
            continue

        latin_greek, english, tentative, abbr, alternatives, example = raw
        if "【" in english or english == "术语对照表":
            continue
        if not english:
            continue

        english, current_prefix = maybe_restore_prefix(english, current_prefix)
        academic = academic_map.get(normalize_term(english), "")
        primary = academic or tentative

        rows.append(
            {
                "latin_greek": latin_greek,
                "term": english,
                "academic_translation": academic,
                "tentative_translation": tentative,
                "abbreviation": abbr,
                "alternative_translation": alternatives,
                "example": example,
                "primary_translation": primary,
            }
        )
    return rows


def build_json(rows):
    out = OrderedDict()
    alt_bucket = {}
    for row in rows:
        term = row["term"]
        if not term:
            continue
        key_base = slugify(term)
        key = key_base
        idx = 2
        while key in out and out[key]["term"] != term:
            key = f"{key_base}_{idx}"
            idx += 1

        translation = row["primary_translation"]
        if key in out:
            # Merge duplicate rows for same term.
            if not out[key]["translation"] and translation:
                out[key]["translation"] = translation
            if not out[key]["abbreviation"] and row["abbreviation"]:
                out[key]["abbreviation"] = row["abbreviation"]
            if row["alternative_translation"]:
                alt_bucket.setdefault(key, set()).add(row["alternative_translation"])
            if (
                translation
                and out[key]["translation"]
                and translation != out[key]["translation"]
                and row["academic_translation"]
            ):
                alt_bucket.setdefault(key, set()).add(translation)
            continue

        explanation_parts = []
        if row["alternative_translation"]:
            explanation_parts.append(f"备选译名：{row['alternative_translation']}")
        explanation = "；".join(explanation_parts)

        out[key] = {
            "term": term,
            "translation": translation,
            "explanation": explanation,
            "abbreviation": row["abbreviation"],
        }

    for key, alts in alt_bucket.items():
        alts = [a for a in sorted(alts) if a and a != out[key]["translation"]]
        if not alts:
            continue
        alt_text = " / ".join(alts)
        if out[key]["explanation"]:
            if alt_text not in out[key]["explanation"]:
                out[key]["explanation"] = f"{out[key]['explanation']}；备选译名：{alt_text}"
        else:
            out[key]["explanation"] = f"备选译名：{alt_text}"

    # Stable sort by term.
    ordered_items = sorted(out.items(), key=lambda kv: kv[1]["term"].lower())
    return OrderedDict(ordered_items)


def write_md(rows):
    lines = [
        "# LTRL 术语对照表（结构化重建）",
        "",
        "> 数据源：`source/reference/疑问汇总及术语对照表.xlsx`（工作表：`综合术语表`）",
        "> 列结构依据源表 `【】` 标注解析；空列按原样保留。",
        "",
        "| 拉丁文/希腊文术语 | 英文术语 | 学术译名 | 术语暂定译名 | 术语简称 | 术语备选译名 | 术语例句 | 主要术语译文 |",
        "|---|---|---|---|---|---|---|---|",
    ]

    for row in rows:
        fields = [
            row["latin_greek"],
            row["term"],
            row["academic_translation"],
            row["tentative_translation"],
            row["abbreviation"],
            row["alternative_translation"],
            row["example"],
            row["primary_translation"],
        ]
        escaped = [f.replace("|", "\\|") for f in fields]
        lines.append("| " + " | ".join(escaped) + " |")

    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb[SHEET]

    academic_map = build_academic_map(ws)
    rows = parse_section1_rows(ws, academic_map)
    data = build_json(rows)

    write_md(rows)
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"section1_rows={len(rows)}")
    print(f"academic_map={len(academic_map)}")
    print(f"json_terms={len(data)}")


if __name__ == "__main__":
    main()
