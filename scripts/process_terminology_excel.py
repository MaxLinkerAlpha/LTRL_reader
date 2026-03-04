#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
处理术语Excel文件
1. 处理带"见"字的译名，将备用译名匹配到主要译名行
2. 删除只包含单个字母的行
3. 生成新的Excel文件
"""

import zipfile
import re
import html as html_module
from pathlib import Path
import shutil

def decode_unicode(text):
    if not text:
        return ""
    return html_module.unescape(text)

def parse_excel(xlsx_path):
    """解析Excel文件"""
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        with z.open('xl/worksheets/sheet1.xml') as f:
            content = f.read().decode('utf-8')
        
        try:
            with z.open('xl/sharedStrings.xml') as f:
                shared = f.read().decode('utf-8')
            shared_strings = re.findall(r'<t[^>]*>(.*?)</t>', shared, re.DOTALL)
            shared_strings = [decode_unicode(s) for s in shared_strings]
        except:
            shared_strings = []
        
        rows = {}
        
        inline_pattern = r'<c r="([A-Z]+)(\d+)"[^>]*t="inlineStr"[^>]*>.*?<t[^>]*>(.*?)</t>.*?</c>'
        for col, row, text in re.findall(inline_pattern, content, re.DOTALL):
            row = int(row)
            text = decode_unicode(text)
            if row not in rows:
                rows[row] = {}
            rows[row][col] = text
        
        s_pattern = r'<c r="([A-Z]+)(\d+)"[^>]*t="s"[^>]*>.*?<v>(\d+)</v>.*?</c>'
        for col, row, idx in re.findall(s_pattern, content, re.DOTALL):
            row = int(row)
            idx = int(idx)
            if idx < len(shared_strings):
                if row not in rows:
                    rows[row] = {}
                rows[row][col] = shared_strings[idx]
    
    return rows, shared_strings

def is_single_letter_row(row_data):
    """检查是否只包含单个字母的行"""
    for col, text in row_data.items():
        text = text.strip()
        if len(text) == 1 and text.isalpha():
            return True
        # 也检查中文单字（如"格"、"数"、"体"、"级"等）
        if len(text) == 1 and '\u4e00' <= text <= '\u9fff':
            return True
    return False

def process_jian_entries(rows):
    """处理带'见'字的条目"""
    # 收集所有带"见"字的条目
    jian_map = {}  # 主要译名 -> 备用译名列表
    rows_to_delete = set()
    
    for row_num, row_data in rows.items():
        if row_num < 4:  # 跳过表头
            continue
        
        c_text = row_data.get('C', '').strip()
        
        # 检查是否包含"见"字
        match = re.match(r'^(.+?)\s*见\s*(.+)$', c_text)
        if match:
            before = match.group(1).strip()
            after = match.group(2).strip()
            
            # 记录需要删除的行
            rows_to_delete.add(row_num)
            
            # 添加到映射
            if after not in jian_map:
                jian_map[after] = []
            jian_map[after].append(before)
    
    print(f"找到 {len(jian_map)} 个主要译名需要匹配")
    print(f"需要删除 {len(rows_to_delete)} 个带'见'字的行")
    
    # 为每个主要译名找到对应的行并添加备用译名
    match_count = 0
    for after, before_list in jian_map.items():
        # 查找包含主要译名的行
        for row_num, row_data in rows.items():
            if row_num < 4:
                continue
            if row_num in rows_to_delete:
                continue
            
            c_text = row_data.get('C', '').strip()
            
            # 检查C列是否完全匹配主要译名
            if c_text == after or c_text.startswith(after + ' ') or c_text.startswith(after + '（'):
                # 找到匹配，添加备用译名到E列
                existing_alt = row_data.get('E', '').strip()
                new_alt = '; '.join(before_list)
                
                if existing_alt:
                    row_data['E'] = existing_alt + '; ' + new_alt
                else:
                    row_data['E'] = new_alt
                
                match_count += 1
                break
    
    print(f"成功匹配 {match_count} 个主要译名")
    
    return rows_to_delete

def remove_single_letter_rows(rows):
    """删除只包含单个字母的行"""
    rows_to_delete = set()
    
    for row_num, row_data in rows.items():
        if row_num < 4:
            continue
        
        if is_single_letter_row(row_data):
            rows_to_delete.add(row_num)
    
    print(f"需要删除 {len(rows_to_delete)} 个单字母行")
    return rows_to_delete

def create_new_excel(rows, rows_to_delete, output_path):
    """创建新的Excel文件"""
    # 创建临时目录
    temp_dir = Path('temp_excel')
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir()
    
    # 解压原始Excel
    with zipfile.ZipFile("source/reference/疑问汇总及术语对照表.xlsx", 'r') as z:
        z.extractall(temp_dir)
    
    # 读取sheet1.xml
    sheet_path = temp_dir / 'xl' / 'worksheets' / 'sheet1.xml'
    with open(sheet_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 构建新的sheet数据
    # 保留前3行（表头）
    max_row = max(rows.keys())
    
    # 构建新的sheetData
    new_sheet_data = '<sheetData>'
    
    for row_num in sorted(rows.keys()):
        if row_num < 4:
            # 保留表头行
            pass
        elif row_num in rows_to_delete:
            # 跳过要删除的行
            continue
        
        if row_num > max_row:
            break
        
        # 这一行会被保留，但需要检查是否有更新的数据
        # 这里我们直接保留原始XML，因为修改XML比较复杂
        # 更简单的做法是重新生成整个sheet
    
    new_sheet_data += '</sheetData>'
    
    # 由于直接修改XML比较复杂，我们采用更简单的方法：
    # 标记要删除的行，在生成MD时跳过
    
    print(f"处理完成，标记了 {len(rows_to_delete)} 行需要跳过")
    
    # 清理
    shutil.rmtree(temp_dir)

def main():
    print("=" * 60)
    print("处理术语Excel文件")
    print("=" * 60)
    
    xlsx_path = Path("source/reference/疑问汇总及术语对照表.xlsx")
    
    print("\n1. 解析Excel...")
    rows, shared_strings = parse_excel(xlsx_path)
    print(f"   共 {len(rows)} 行")
    
    print("\n2. 处理带'见'字的条目...")
    jian_rows = process_jian_entries(rows)
    
    print("\n3. 查找单字母行...")
    single_letter_rows = remove_single_letter_rows(rows)
    
    # 合并要删除的行
    all_rows_to_delete = jian_rows.union(single_letter_rows)
    print(f"\n总共需要删除 {len(all_rows_to_delete)} 行")
    
    print("\n4. 生成处理后的数据...")
    # 生成一个处理后的JSON文件，便于后续使用
    processed_data = []
    
    for row_num in sorted(rows.keys()):
        if row_num < 4:
            continue
        if row_num in all_rows_to_delete:
            continue
        
        row_data = rows[row_num]
        
        # 收集数据
        entry = {
            'row': row_num,
            'latin': row_data.get('A', ''),
            'english': row_data.get('B', ''),
            'translation': row_data.get('C', ''),
            'abbreviation': row_data.get('D', ''),
            'alternatives': row_data.get('E', ''),
            'example': row_data.get('F', '')
        }
        
        processed_data.append(entry)
    
    print(f"   保留 {len(processed_data)} 行数据")
    
    # 保存为JSON
    import json
    with open('temp_processed_terminology.json', 'w', encoding='utf-8') as f:
        json.dump(processed_data, f, ensure_ascii=False, indent=2)
    print(f"   ✓ 已保存: temp_processed_terminology.json")
    
    print("\n5. 生成新的Excel文件...")
    # 这里我们需要使用openpyxl来创建新的Excel文件
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "术语对照表"
        
        # 写入表头（第1-3行）
        ws['A1'] = '参考术语源：https://zhuanlan.zhihu.com/p/653809137'
        ws['A2'] = '术语对照表'
        ws['A3'] = '【拉丁文/希腊文术语】'
        ws['B3'] = '【英文术语】'
        ws['C3'] = '【术语暂定译名】'
        ws['D3'] = '【术语简称】'
        ws['E3'] = '【术语备选译名】'
        ws['F3'] = '【术语例句】'
        
        # 写入数据
        for idx, entry in enumerate(processed_data, start=4):
            ws[f'A{idx}'] = entry['latin']
            ws[f'B{idx}'] = entry['english']
            ws[f'C{idx}'] = entry['translation']
            ws[f'D{idx}'] = entry['abbreviation']
            ws[f'E{idx}'] = entry['alternatives']
            ws[f'F{idx}'] = entry['example']
        
        # 保存
        output_path = Path("source/reference/疑问汇总及术语对照表_已处理.xlsx")
        wb.save(output_path)
        print(f"   ✓ 已保存: {output_path}")
        print(f"   共 {len(processed_data) + 3} 行（含表头）")
        
    except ImportError:
        print("   ✗ 未安装openpyxl，无法生成Excel文件")
        print("   请运行: pip install openpyxl")
        print(f"   已保存JSON文件作为替代: temp_processed_terminology.json")

if __name__ == '__main__':
    main()
